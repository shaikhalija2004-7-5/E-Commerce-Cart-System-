# from openpyxl import load_workbook
# wb = load_workbook('goods list.xlsx')
# sheet = wb.active
#
# for row in sheet.iter_rows(values_only=True):
#     print(row)
# from curses.ascii import isdigit

print('--------------------------🛒 E-Commerce Cart System 🛒---------------------------')
print('--------Welcome to our Store---------')
#-------------------------------INFO
class information():
    def __init__(self):
        while True:
            self.Name = input('Please enter your name: ').title()
            if not self.Name  or self.Name.isalpha()== False:
                print('Please enter a valid name. Name must be alphabet')
                continue
            else:
                print(f'The customer name is :- {self.Name}')
                break

        while True:
            self.EMail = input('Please enter your E-Mail: ').lower()
            if not self.EMail or not self.EMail.endswith('@gmail.com') :
                print('Please enter a valid E-Mail.!')
                continue
            else:
                print(f'The E-Mail is :- {self.EMail}\n')
                break

#------------------------GOODS lIST
class GOODS():
    def __init__(self):
        from openpyxl import load_workbook
        self.WB = load_workbook('goods list.xlsx')
        self.WS = self.WB.active

    def show_goods(self):
        while True:
            self.Goods = input('Do you wanna check list of goods? (y/n): \n').upper()
            if self.Goods == 'Y':
                print('-----------AVAILBLE GOODS IN STOCK -------------')
                for row in self.WS.iter_rows(values_only=True):
                    print(row)
                break
            else :
                print('ok! NO problem')

#----------------------------- ORDERS LIST
class orders(GOODS):
    def __init__(self):
        super().__init__()
        self.carts =[]

    def order(self):
        while True:
            while True:
                try:
                    productid = int(input('\nPlease enter product id: '))
                    break  # valid input
                except ValueError:
                    print('❌ Please enter a valid numeric product ID!')
            found = False
            for row in self.WS.iter_rows():
                if productid == row[0].value :
                    print(f'yes the product is availble ! ')
                    print([cell.value for cell in row])
                    found = True

                    Quantity = int(input('\nPlease enter quantity: '))
                    Current_stock = row[3].value
                    if Quantity <=  Current_stock :
                        print(f'yes the stock is availble ! ')
                        row[3].value = Current_stock - Quantity
                        data = [cell.value for cell in row]
                        print(f'\nProduct add to your cart 🛒 ! {Quantity}')
                        data.append(Quantity)
                        self.carts.append(data)
                        self.WB.save('goods list.xlsx')
                        break

                    else:
                        print("❌ Not enough stock!")
                    break
            if not found:
                print('Sorry ! for unavailability 🙏!')

            othergoods = input("\nDo you want other items? (y/n): ").upper()

            if othergoods != 'Y':
                break

        print('\n🛒 Your Cart:')
        for item in self.carts:
            print(item)

#-----------------------BILLING
class billing():
    def __init__(self, name, email,carts):
        self.Name = name
        self.EMail = email
        self.carts = carts


    def bills(self):
        print('-\n----------------BILL----------')
        print('----------MOMS KITCHEN GENERAL STORE--------------')
        print('-------------📍SAI Apt. Dadar(w)-401209---------------')
        print('---------for any contactand queries:- ☎️ +91 9813451550 / 5896321458----------')
        print('Thankyou for visiting our store 🙏\n')
        print(f'Name of Customer :- {self.Name}')
        print(f'Email-id :- {self.EMail}\n')
        total_quantity = 0
        total_amount = 0
        print(f'Your Product is :- ')
        print('ID | Name | Category | Qty | Rate | Total')

        for item in self.carts:
            product_id = item[0]
            Name = item[1]
            Category = item[2]
            Qty = item[5]
            Rate = item[4]


            Total = Qty * Rate
            total_quantity += Qty
            total_amount += Total
            print(f'{product_id} | {Name} | {Category} | {Qty} | {Rate} | {Total}')

        print(f'\nTotal Quantity Purchased | {total_quantity}')
        print(f'Total Amount Of Purchased | ₹{total_amount} 💸\n')
        self.Cgst = total_amount *0.07
        self.Sgst = total_amount *0.07
        print(f'Cgst amount | ₹{(self.Cgst).__round__(2)} 💸\nSgst amount | ₹{(self.Sgst).__round__(2)} 💸')
        print(f'Total Gst amount | ₹{(self.Sgst+self.Cgst).__round__(2)} 💸\n')
        discount = 0
        if total_amount >2500:
            discount = total_amount *0.5
            print(f'Allowed Discount above >5000 shopping | ₹{round(discount,2)} 💸')
        else :
            print(f'NO discount allowed')
        print(f'The Bill Amount is | ₹{total_amount+self.Cgst+self.Sgst-discount} 💸\n')
        print('\n Thankyou For Shopping!')


# ----------------------call functions
A=information()
G = GOODS()
G.show_goods()
C = orders()
C.order()
D = billing(A.Name,A.EMail,C.carts)
D.bills()